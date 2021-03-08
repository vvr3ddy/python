from openpyxl import load_workbook
from openpyxl import Workbook
# The source xlsx file is named as sample.xlsx
wb=load_workbook("sample.xlsx")

ws = wb.active
first_column = ws['A']
raw_bin = []
# store the contents in a list
for x in range(len(first_column)): 
    raw_bin.append(str(first_column[x].value)) 


#print the contents
print(raw_bin)

#split size = 3 change accordingly
n = 3
processed_bin = []

#split the value into 3 bit each and print
for elem in raw_bin:
    
    curVal = elem
    tmp_bin = [(curVal[i:i+n]) for i in range(0, len(curVal), n)] 
    processed_bin.append(tmp_bin)
for i in range(len(processed_bin)):
    print(processed_bin[i])

saveBook = Workbook()
filepath = "Processed.xlsx"
sheet=wb.active
for row in processed_bin:
    sheet.append(row)
wb.save(filepath)
