from openpyxl import load_workbook
from openpyxl import Workbook
from pprint import pprint

# The source xlsx file is named as sample.xlsx
wb=load_workbook("sample.xlsx")
#destination is saved as results.xlsx
filepath = "results.xlsx"

#define the current sheet active in the workbook
currentSheet = wb.active
#define the target workbook sheet where we save our result and name it results


rowData = []
for row in currentSheet.values:
    for value in row:
        if(value!=None):
            print(value)
            rowData.append(value)
print(rowData) # all the data has been now fetched into a 1d list
targetSheet = wb.create_sheet("Results",0)
for value in rowData:
    print(value)
    targetSheet.append([value])

wb.save(filepath)
